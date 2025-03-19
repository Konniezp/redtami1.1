import base64
from datetime import datetime, date
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
from io import BytesIO
import requests
import numpy as np

from django.contrib import auth, messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import connection
from django.db.models import Count, F, Max
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.utils import timezone


from openpyxl import Workbook

from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import viewsets
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.permissions import IsAdminUser
from rest_framework.parsers import JSONParser
from django.core.exceptions import ObjectDoesNotExist

from .forms import *
from .models import *
from .serializer import *

from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from django.views.decorators.csrf import csrf_exempt
import json
import locale 

import logging

locale.setlocale(locale.LC_TIME, 'es_ES')
    
@login_required
def home(request):
    return render(request, "home.html")


def login(request):
    if request.method == "POST":
        username = request.POST["username"]
        password = request.POST["password"]

        user = auth.authenticate(username=username, password=password)

        if user is not None:
            auth.login(request, user)
            return redirect("home")
        else:
            return render(request, "registration/login.html")

    return render(request, "registration/login.html")

@login_required
def logout(request):
    auth.logout(request)


# --------------------- Respuestas de Usuario --------------------- #

#Home
@login_required
def respuestasHome(request):
    return render(request, "respuestas/respuestasHome.html")

@login_required
def opcVisualFRM(request):
    return render(request, "respuestas/opcVisualFRM.html")

@login_required
def opcVisualFRNM(request):
    return render(request, "respuestas/opcVisualFRNM.html")

@login_required
def opcVisualDS(request):
    return render(request, "respuestas/opcVisualDS.html")

# Base de datos
@login_required
def datosPerfil(request):
    Datos = Usuario.objects.all().order_by("-Fecha_Ingreso")
    data = {
        "Datos": Datos,
    }
    return render(request, "respuestas/datosPerfil.html", data)

@login_required
def datosPreguntas(request):
    Datos = RespUsuarioTamizaje.objects.select_related(
        "respuesta_TM", "respuesta_TM__id_pregunta").values("id_manychat",
        "respuesta_TM__id_pregunta__pregunta", "respuesta_TM__opc_respuesta_TM",
        "fecha_respuesta").order_by("-fecha_respuesta")
    data = {
        "Datos": Datos,
    }
    return render(request, "respuestas/datosPreguntas.html", data)

@login_required
def datosListadoOrdenado(request):

    if request.method == "POST":
            password_ingresada = request.POST.get("password")
            if password_ingresada == settings.ACCESO_LISTADO:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT DISTINCT us.id_manychat,
                        Whatsapp, edad, opc.opc_respuesta_TM AS Antecedentes, 
                        COALESCE(ult.tiempo_transc_ult_mamografia, 1000) AS Ult_mamografia
                        FROM botApp_opctamizaje opc 
                        LEFT JOIN botApp_respusuariotamizaje rt ON opc.id = rt.respuesta_TM_id
                        LEFT JOIN botApp_usuario us ON us.id_manychat = rt.id_manychat
                        LEFT JOIN botApp_ultima_mamografia_anio ult ON us.id_manychat = ult.id_manychat_id
                        WHERE respuesta_TM_id IN (1,2,3)         
                        ORDER BY ult.tiempo_transc_ult_mamografia DESC;
                        """)
                    columns = [col[0] for col in cursor.description]
                    datos = cursor.fetchall()

                #Descifra datos 
                datos_descifrados=[]
                for row in datos:
                    id_manychat, Whatsapp, edad, antecedentes, ultima_mamografia = row
                
                    #Intenta descifrar cada campo encriptado
                    try:
                        Whatsapp_descifrado = decrypt_data(Whatsapp) if Whatsapp else "No disponible"
                    except:
                        Whatsapp_descifrado = "Error al descifrar Whatsapp"
                    
                    datos_descifrados.append({
                        "id": id_manychat,
                        "Whatsapp": Whatsapp_descifrado,
                        "edad": edad,
                        "Antecedentes_familiares": antecedentes,
                        "Ult_mamografia": ultima_mamografia,
                    })

                return render(request, "respuestas/datosListadoOrdenado.html", {"Datos": datos_descifrados})
            else:
                error = "Contraseña incorrecta"
    else:
        error = None
        
    return render(request, "respuestas/form_contrasena_listado.html", {"error": error})
#Ajustar anchos de columnas según el largo de la celda

def ajustar_ancho_columnas(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell.value: 
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        ws.column_dimensions[col_letter].width = max_length + 2  

# Aplica color a la primera fila de encabezados 
def background_colors(ws):
    color = "00a0b3a8" 
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    for cell in ws[1]: 
        cell.fill = fill

def crear_excel_desde_db():
    wb = Workbook()

    # Hoja 1: Respuestas Usuario
    ws_respuestas_usuario = wb.active
    ws_respuestas_usuario.title = 'Respuestas Usuario'

    preguntas = PregTamizaje.objects.all()
    lista_preguntas = [ "id_manychat"] + [pregunta.pregunta for pregunta in preguntas]
    ws_respuestas_usuario.append(lista_preguntas)

    usuarios_respuestas = RespUsuarioTamizaje.objects.select_related('respuesta_TM', 'respuesta_TM__id_pregunta').values(
        'id_manychat', 
        'respuesta_TM__id_pregunta__pregunta', 
        'respuesta_TM__opc_respuesta_TM'
    )

    dict_respuestas = {}

    for respuesta in usuarios_respuestas:
        id_manychat = respuesta[ 'id_manychat']
        pregunta = respuesta['respuesta_TM__id_pregunta__pregunta']
        respuesta_usuario = respuesta['respuesta_TM__opc_respuesta_TM']
        if id_manychat not in dict_respuestas:
            dict_respuestas[id_manychat] = {}
        dict_respuestas[id_manychat][pregunta] = respuesta_usuario

    for id_manychat, respuestas_usuario in dict_respuestas.items():
        fila = [id_manychat]
        for pregunta in preguntas:
            respuesta = respuestas_usuario.get(pregunta.pregunta, '')
            fila.append(respuesta)
        ws_respuestas_usuario.append(fila)

    # Ajustar ancho de columnas y color de fondo 
    ajustar_ancho_columnas(ws_respuestas_usuario)
    background_colors(ws_respuestas_usuario)

    # Hoja 2: Datos Perfil
    ws_datos_perfil = wb.create_sheet(title='Datos Perfil')
    campos_usuario = [field.name for field in Usuario._meta.fields if field.name not in ['Comuna_Usuario']]
    ws_datos_perfil.append(campos_usuario)

    for usuario in Usuario.objects.all():
        datos_usuario = [str(getattr(usuario, campo)) for campo in campos_usuario]
        ws_datos_perfil.append(datos_usuario)

    # Ajustar ancho de columnas y color de fondo 
    ajustar_ancho_columnas(ws_datos_perfil)
    background_colors(ws_datos_perfil)

    # Guardar el archivo
    nombre_archivo = 'reporte_respuestas.xlsx'
    wb.save(nombre_archivo)

    return nombre_archivo

def descargar_excel(request):
    # Llama a la función para crear el Excel
    nombre_archivo = crear_excel_desde_db()

    # Abre el archivo Excel y lo envía como una respuesta de descarga
    with open(nombre_archivo, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'
        return response

def crear_excel_listado_ordenable(request):
    with connection.cursor() as cursor:
        cursor.execute("""
        SELECT 
        us.id_manychat, Whatsapp, edad, opc.opc_respuesta_TM AS Antecedentes, 
        COALESCE(ult.tiempo_transc_ult_mamografia, 1000) AS Ult_mamografia
        FROM botApp_opctamizaje opc 
        LEFT JOIN botApp_respusuariotamizaje rt ON opc.id = rt.respuesta_TM_id
        LEFT JOIN botApp_usuario us ON us.id_manychat = rt.id_manychat
        LEFT JOIN botApp_ultima_mamografia_anio ult ON us.id_manychat = ult.id_manychat_id         
        WHERE rt.respuesta_TM_id IN (25, 26, 27)
        ORDER BY ult.tiempo_transc_ult_mamografia DESC;
        """)
        columns = [col[0] for col in cursor.description]
        data = cursor.fetchall()

    # Crear un archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado Ordenable"

    # Agregar encabezados
    ws.append(columns)

    # Agregar los datos fila por fila
    datos_descifrados = []
    for row in data:
        id_manychat, Whatsapp, edad, antecedentes, ultima_mamografia = row

        # Intenta descifrar los datos
        try:
            Whatsapp_descifrado = decrypt_data(Whatsapp) if Whatsapp else "No disponible"
        except:
            Whatsapp_descifrado = "Error al descifrar Whatsapp"

        # Agregar los datos descifrados a la lista y a la hoja de Excel
        fila_descifrada = (id_manychat, Whatsapp_descifrado, edad, antecedentes, ultima_mamografia)
        datos_descifrados.append(fila_descifrada)
        ws.append(fila_descifrada) 

    # Ajustar ancho de columnas y color de fondo 
    ajustar_ancho_columnas(ws)
    background_colors(ws)

    # Preparar la respuesta HTTP
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="ListadoOrdenable.xlsx"'

    # Guardar el archivo en la respuesta HTTP
    wb.save(response)
    return response

def crear_excel_datos_preguntas(resquest):
    wb = Workbook()
    ws_datos_preg = wb.active
    ws_datos_preg.title = "Preguntas generales"

    preguntas = PregTamizaje.objects.all()
    lista_preguntas = ['id_manychat', 'Preguntas', 'Respuestas', 'Fecha Respuesta'] 
    ws_datos_preg.append(lista_preguntas)

    usuarios_respuestas = RespUsuarioTamizaje.objects.select_related(
        "respuesta_TM", "respuesta_TM__id_pregunta").values("id_manychat",
        "respuesta_TM__id_pregunta__pregunta", "respuesta_TM__opc_respuesta_TM",
        "fecha_respuesta").order_by("-fecha_respuesta")
    

    for respuesta in usuarios_respuestas:

        pregunta = respuesta['respuesta_TM__id_pregunta__pregunta']
        respuesta_usuario = respuesta['respuesta_TM__opc_respuesta_TM']
        fecha_respuesta = respuesta['fecha_respuesta'].replace(tzinfo=None) if respuesta['fecha_respuesta'] else ''
        fila = [
            respuesta['id_manychat'],
            pregunta,  
            respuesta_usuario, 
            fecha_respuesta,  
        ]
        ws_datos_preg.append(fila)
    
    # Ajustar ancho de columnas y color de fondo 
    ajustar_ancho_columnas(ws_datos_preg)
    background_colors(ws_datos_preg)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Datos preguntas.xlsx"'

    # Guardar el archivo en la respuesta HTTP
    wb.save(response)
    return response

# --------------------- Reporteria --------------------- #

# Configuración global para fuentes de gráficos
plt.rcParams['font.family'] = 'sans-serif'  
plt.rcParams['font.sans-serif'] = 'Calibri' 
plt.rcParams['font.size'] = 11
plt.rcParams['axes.titlesize'] =20
plt.rcParams['axes.labelsize']= 13
plt.rcParams['axes.labelpad']=10

def generar_grafico_usuario_por_edad():

    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT edad, COUNT(*) 
            FROM botApp_usuario 
            GROUP BY edad 
            ORDER BY edad;"""
        )
        resultados = cursor.fetchall()

    edades = []
    cantidades = []

    for resultado in resultados:
        edad, cantidad = resultado
        edades.append(edad)
        cantidades.append(cantidad)
    
    plt.figure(figsize=[18, 8])
    plt.bar(edades, cantidades, color="#79addc")
    plt.xlabel("Edad")
    plt.ylabel("Número de Usuarias")
    plt.title("Usuarias por edad", pad=20)
    plt.xticks(range(min(edades), max(edades) + 1, 1))
    

    # Agregar etiquetas en las barras
    for edad, cantidad in zip(edades, cantidades):
        plt.text(edad, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_anio_nacimiento():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT YEAR(AnioNacimiento) as anio, COUNT(*) 
            FROM botApp_usuario 
            GROUP BY YEAR(AnioNacimiento) ORDER BY anio ASC;"""
        )
        resultados = cursor.fetchall()

    anios = []
    cantidades = []

    for resultado in resultados:
        anio, cantidad = resultado
        anios.append(anio)
        cantidades.append(cantidad)
    
    plt.figure(figsize=[18, 8])
    plt.bar(anios, cantidades, color="#79addc")
    plt.xlabel("Año de Nacimiento")
    plt.ylabel("Número de Usuarios")
    plt.title("Usuarias por Año de Nacimiento", pad=20)
    plt.xticks(range(min(anios), max(anios)+1,1), rotation = 90)

    # Agregar etiquetas en las barras
    for anio, cantidad in zip(anios, cantidades):
        plt.text(anio, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_respuestas_por_dia():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT DATE(Fecha_Ingreso), COUNT(*) 
            FROM botApp_usuario 
            GROUP BY DATE(Fecha_Ingreso);  """
        )
        resultados = cursor.fetchall()

    fechas = []
    cantidades = []

    for resultado in resultados:
        fecha, cantidad = resultado
        fechas.append(datetime.strftime(fecha, "%d-%m-%Y"))
        cantidades.append(cantidad)

    plt.plot(fechas, cantidades, marker="o", linestyle="-", color="#79addc")
    plt.xlabel("Fecha de Respuesta")
    plt.ylabel("Número de Respuestas")
    plt.title("Respuestas por Día", pad=20)
    plt.xticks(rotation = 90)
    plt.tight_layout() 
    
    # Agregar los valores de cada punto
    for fecha, cantidad in zip(fechas, cantidades):
        plt.annotate(f"{cantidad}", (fecha, cantidad), textcoords="offset points", xytext=(0,10), ha='center')

    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_ingresos_por_comuna():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT c.nombre_comuna, COUNT(*) AS TotalIngresos 
            FROM botApp_comuna c JOIN botApp_usuario u ON u.Comuna_Usuario_id = c.cod_comuna 
            GROUP BY c.nombre_comuna;"""
        )
        resultados = cursor.fetchall()

    comunas = [result[0] for result in resultados]
    total_ingresos = [result[1] for result in resultados]

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(total_ingresos, labels=comunas, autopct=lambda pct: f"{pct:.1f}%\n{int(pct/100 * sum(total_ingresos)+ 0.5)} ingresos", startangle=90)
    ax.axis('equal')  # Asegura que el gráfico sea un círculo en lugar de una elipse
    ax.set_title('Distribución de Ingresos por Comuna', pad=20)

    # Ajustar el tamaño de la fuente en los textos
    for text, autotext in zip(texts, autotexts):
        text.set(size=8)
        autotext.set(size=8)

    # Convertir el gráfico a base64
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    plt.close()
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

    return imagen_base64

def generar_grafico_referencias():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT u.Referencia, COUNT(*) AS TotalIngresos 
            FROM botApp_usuario u 
            GROUP BY u.Referencia;"""
        )
        resultados = cursor.fetchall()

    referencias = [result[0] for result in resultados]
    total_ingresos = [result[1] for result in resultados]

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(total_ingresos, labels=referencias, autopct=lambda pct: f"{pct:.1f}%\n{int(pct/100 * sum(total_ingresos))} ingresos", startangle=90)
    ax.axis('equal')  # Asegura que el gráfico sea un círculo en lugar de una elipse
    ax.set_title('Distribución de Ingresos por Referencia', pad=20)

    # Ajustar el tamaño de la fuente en los textos
    for text, autotext in zip(texts, autotexts):
        text.set(size=8)
        autotext.set(size=8)

    # Convertir el gráfico a base64
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    plt.close()
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

    return imagen_base64

def generar_grafico_pregunta1():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT respuesta_TM_id, COUNT(*) 
            FROM botApp_respusuariotamizaje 
            WHERE respuesta_TM_id IN (1, 2, 3) 
            GROUP BY respuesta_TM_id;"""
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = OpcTamizaje.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.opc_respuesta_TM)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.opc_respuesta_TM} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots(figsize=(8, 8))
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['#79addc', '#EFB0C9', '#A5F8CE'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('¿Te has realizado una mamografía?', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_pregunta2():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT respuesta_TM_id, COUNT(*) 
            FROM botApp_respusuariotamizaje 
            WHERE respuesta_TM_id IN (7, 8) 
            GROUP BY respuesta_TM_id;"""
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = OpcTamizaje.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.opc_respuesta_TM)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.opc_respuesta_TM} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots(figsize=(8, 8))
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['#79addc', '#EFB0C9', '#DEB3EB'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('¿Recuerdas cuando fue tu última mamografía?', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_pregunta3():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT respuesta_TM_id, COUNT(*) 
            FROM botApp_respusuariotamizaje u 
            WHERE respuesta_TM_id IN (10,11,12) 
            GROUP BY respuesta_TM_id;"""
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = OpcTamizaje.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.opc_respuesta_TM)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.opc_respuesta_TM} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots(figsize=(8, 8))
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['#DE8F5F', '#FFB26F', '#FFB38E', '#DEB3EB'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('Fecha de la última mamografía', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_pregunta4():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT respuesta_TM_id, COUNT(*) 
            FROM botApp_respusuariotamizaje u 
            WHERE respuesta_TM_id IN (14, 15, 16) 
            GROUP BY respuesta_TM_id;"""
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = OpcTamizaje.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.opc_respuesta_TM)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.opc_respuesta_TM} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots(figsize=(8, 8))
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['#79addc', '#EFB0C9','#A5F8CE', '#DEB3EB'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('¿Tienes los archivos e informe \nde tu última mamografía?', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_pregunta5():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT respuesta_TM_id, COUNT(*) 
            FROM botApp_respusuariotamizaje u
            WHERE respuesta_TM_id IN (22, 23, 24) 
            GROUP BY respuesta_TM_id;"""
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = OpcTamizaje.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.opc_respuesta_TM)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.opc_respuesta_TM} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots(figsize=(8, 8))
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['#79addc', '#EFB0C9', '#A5F8CE'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('¿Te gustaría recibir más información \nsobre el cuidado y prevención del cáncer de mama?', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_pregunta6():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT respuesta_TM_id, COUNT(*) 
            FROM botApp_respusuariotamizaje
            WHERE respuesta_TM_id IN (25, 26, 27)
            GROUP BY respuesta_TM_id;"""
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = OpcTamizaje.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.opc_respuesta_TM)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.opc_respuesta_TM} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots(figsize=(8, 8))
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['#79addc', '#EFB0C9', '#A5F8CE'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('¿Tienes un familiar directo con cáncer de mama?\n(hermana, mama, tía, abuela)', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64


def generar_grafico_mamografia_si_por_edad():

    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT us.edad, COUNT(*) as Cantidad 
            FROM botApp_respusuariotamizaje ur JOIN botApp_usuario us ON ur.id_manychat = us.id_manychat 
            WHERE respuesta_TM_id IN (1)
            GROUP BY us.edad ORDER BY edad ASC;""" 
        )
        resultados = cursor.fetchall()

    edades = []
    cantidades = []

    for resultado in resultados:
        edad, cantidad = resultado
        edades.append(edad)
        cantidades.append(cantidad)

    if not edades: 
        return None
    
    plt.figure(figsize=[18, 8])
    plt.bar(edades, cantidades, color="#79addc")
    plt.xlabel("Edad")
    plt.ylabel("Número de Usuarias")
    plt.title("Mamografías por edad Respuesta Si", pad=20)
    plt.xticks(range(min(edades), max(edades) + 1, 1))
    plt.yticks(range(0,11,1))

    # Agregar etiquetas en las barras
    for edad, cantidad in zip(edades, cantidades):
        plt.text(edad, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_mamografia_no_por_edad():

    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT us.edad, COUNT(*) as Cantidad 
            FROM botApp_respusuariotamizaje ur JOIN botApp_usuario us ON ur.id_manychat = us.id_manychat
            WHERE respuesta_TM_id IN (2)
            GROUP BY edad ORDER BY edad ASC;"""
        )
        resultados = cursor.fetchall()

    edades = []
    cantidades = []

    for resultado in resultados:
        edad, cantidad = resultado
        edades.append(edad)
        cantidades.append(cantidad)

    plt.figure(figsize=[18, 8])
    plt.bar(edades, cantidades, color="#EFB0C9")
    plt.xlabel("Edad")
    plt.ylabel("Número de Usuarias")
    plt.title("Mamografías por edad Respuesta No", pad=20)
    plt.xticks(range(min(edades), max(edades) + 1, 1))
 
    # Agregar etiquetas en las barras
    for edad, cantidad in zip(edades, cantidades):
        plt.text(edad, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def mamografia_por_edad_si_no():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT us.edad, COUNT(*) as Cantidad, ur.respuesta_TM_id
            FROM botApp_respusuariotamizaje ur 
            JOIN botApp_usuario us ON ur.id_manychat = us.id_manychat 
            WHERE respuesta_TM_id IN (1,2) 
            GROUP BY edad, ur.respuesta_TM_id 
            ORDER BY edad ASC;"""         
        )
        resultados = cursor.fetchall()

    edades = []
    cantidades_si = []
    cantidades_no = []
    cantidades_no_recuerdo = []

    # Iteramos sobre los resultados
    for resultado in resultados:
        edad, cantidad, respuesta = resultado

        # Si la edad no está en la lista, la agregamos con inicialización de cantidades
        if edad not in edades:
            edades.append(edad)
            cantidades_si.append(0)  
            cantidades_no.append(0)  
            cantidades_no_recuerdo.append(0)

        # Obtenemos el índice correspondiente a la edad
        index = edades.index(edad)

        if respuesta == 1:
            cantidades_si[index] += cantidad
        elif respuesta == 2: 
            cantidades_no[index] += cantidad
        elif respuesta == 3: 
            cantidades_no_recuerdo[index] += cantidad
        

    # Crear el gráfico
    plt.figure(figsize=[18, 8])
    plt.bar(edades, cantidades_si, color="#79addc", label="Cantidad Sí")
    plt.bar(edades, cantidades_no, color="#EFB0C9", bottom=cantidades_si, label="Cantidad No")
    plt.bar(edades, cantidades_no_recuerdo, color="#A5F8CE", bottom=np.array(cantidades_si) + np.array(cantidades_no), label="Cantidad No recuerdo")
    plt.xlabel("Edad")
    plt.ylabel("Número de Usuarias")
    plt.title("Mamografías por Edad", pad=20)
    plt.xticks(np.arange(min(edades), max(edades) + 1, 1)) 
    plt.legend()

    for edad, cantidad_si, cantidad_no, cantidad_no_recuerdo in zip(edades, cantidades_si, cantidades_no, cantidades_no_recuerdo):
        if cantidad_si > 0:
            plt.text(edad, cantidad_si / 2, str(cantidad_si), ha='center', va='bottom', color='black')
        if cantidad_no > 0:
            plt.text(edad, cantidad_si + cantidad_no / 2, str(cantidad_no), ha='center', va='bottom', color='black')
        if cantidad_no_recuerdo > 0:
            plt.text(edad, cantidad_si + cantidad_no + cantidad_no_recuerdo / 2, str(cantidad_no_recuerdo), 
                     ha='center', va='center', color='black', fontsize=10, fontweight='bold')
            
   # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_tiempo_trascurrido():

    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT tiempo_transc_ult_mamografia, COUNT(*) 
            FROM botApp_ultima_mamografia_anio u 
            GROUP BY tiempo_transc_ult_mamografia ORDER BY tiempo_transc_ult_mamografia ASC;"""
        )
        resultados = cursor.fetchall()

    rango_uno_etiqueta = "1"
    rango_dos_etiqueta = "2"
    rango_tres_etiqueta = "Más de 3"

    opciones_anios = [rango_uno_etiqueta, rango_dos_etiqueta, rango_tres_etiqueta]
    cantidades = [0, 0, 0]

    for resultado in resultados:
        anio, cantidad = resultado
        if anio == 1:
            cantidades[0] += cantidad
        elif anio == 2:
            cantidades[1] += cantidad
        elif anio == 3:
            cantidades[2] += cantidad

    plt.figure(figsize=[18, 8])
    plt.bar(opciones_anios, cantidades, color="#79addc")
    plt.xlabel("Años transcurridos")
    plt.ylabel("Cantidad de usuarias")
    plt.title("Tiempo transcurrido desde última mamografía", pad=20)

    # Agregar etiquetas en las barras
    for i, cantidad in enumerate(cantidades):
        plt.text(i, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_por_rango_edad():

    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT edad, COUNT(*) 
            FROM botApp_usuario u
            GROUP BY edad ORDER BY edad;"""  
        )
        resultados = cursor.fetchall()

    r_uno_edad_eti = "Menor de 50 años"
    r_dos_edad_eti = "Entre 50 y 69 años"
    r_tres_edad_eti = "Mayor a 69 años"
    opciones_edad = [r_uno_edad_eti, r_dos_edad_eti, r_tres_edad_eti]
    cantidades = [0, 0, 0]

    edad_min = 50
    edad_max = 69

    for resultado in resultados:
        edad, cantidad = resultado
        if edad <edad_min:
            cantidades[0] += cantidad
        elif edad >= edad_min and edad <= edad_max:
            cantidades[1] += cantidad
        elif edad > edad_max :
            cantidades[2] += cantidad

    plt.figure(figsize=[18, 8])
    plt.bar(opciones_edad, cantidades, color=['#DE8F5F', '#FFB26F', '#FFB38E'])
    plt.xlabel("Rango de edad según guía clínica")
    plt.ylabel("Cantidad de usuarias")
    plt.title("Cantidad de usuarias por rango de edad", pad=20)

    # Agregar etiquetas en las barras
    for i, cantidad in enumerate(cantidades):
        plt.text(i, cantidad, str(cantidad), ha='center', va='bottom')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def mamografia_por_edad_si_no_rango_edad():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT us.edad, COUNT(*) as Cantidad, ur.respuesta_TM_id
            FROM botApp_respusuariotamizaje ur 
            JOIN botApp_usuario us ON ur.id_manychat = us.id_manychat 
            WHERE respuesta_TM_id IN (1,2,3)
            GROUP BY edad, ur.respuesta_TM_id 
            ORDER BY edad ASC;"""  
        )
        resultados = cursor.fetchall()

    r_uno_edad_eti = "Menor de 50 años"
    r_dos_edad_eti = "Entre 50 y 69 años"
    r_tres_edad_eti = "Mayor a 69 años"
    opciones_anios = [r_uno_edad_eti, r_dos_edad_eti, r_tres_edad_eti]
    cantidades_si = [0, 0, 0]
    cantidades_no = [0, 0, 0]
    cantidades_no_recuerdo = [0,0,0]

    edad_min = 50
    edad_max = 69

    # Iteramos sobre los resultados
    for edad, cantidad, respuesta in resultados:
        index = 0 if edad < edad_min else (1 if edad <= edad_max else 2)
        
        if respuesta == 1:
            cantidades_si[index] += cantidad
        elif respuesta == 2:
            cantidades_no[index] += cantidad
        elif respuesta == 3:
            cantidades_no_recuerdo[index] += cantidad
    
    # Crear el gráfico
    plt.figure(figsize=[18, 8])
    plt.bar(opciones_anios, cantidades_si, color="#79addc", label="Cantidad Si")
    plt.bar(opciones_anios, cantidades_no, color="#EFB0C9", bottom=cantidades_si, label="Cantidad No")
    plt.bar(opciones_anios, cantidades_no_recuerdo, color="#A5F8CE", bottom=np.array(cantidades_si) + np.array(cantidades_no), label="Cantidad No recuerdo")
    plt.xlabel("Rango de edad según guía clínica")
    plt.ylabel("Número de Usuarias")
    plt.title("Mamografía por rango de edad", pad=20)
    plt.legend()
    # Agregar etiquetas para las barras de cantidades_si
    for i, (edad, cantidad_si) in enumerate(zip(opciones_anios, cantidades_si)):
        if cantidad_si > 0:
            plt.text(i, cantidad_si / 2, str(cantidad_si), ha='center', va='center', color='black', fontsize=10, fontweight='bold')

    # Agregar etiquetas para las barras de cantidades_no
    for i, (edad, cantidad_si, cantidad_no) in enumerate(zip(opciones_anios, cantidades_si, cantidades_no)):
        if cantidad_no > 0:
            plt.text(i, cantidad_si + cantidad_no / 2, str(cantidad_no), ha='center', va='center', color='black', fontsize=10, fontweight='bold')

    # Agregar etiquetas para la barra de cantidades_no_recuerdo
    for i, (edad, cantidad_si, cantidad_no, cantidad_no_recuerdo) in enumerate(zip(opciones_anios, cantidades_si, cantidades_no, cantidades_no_recuerdo)):
        if cantidad_no_recuerdo > 0:
            plt.text(i, cantidad_si + cantidad_no + cantidad_no_recuerdo / 2, str(cantidad_no_recuerdo), ha='center', va='center', color='black', fontsize=10, fontweight='bold')

   # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64
    

def mamografia_por_edad_si_no_rango_edad_agrupado():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT us.edad, COUNT(*) as Cantidad, ur.respuesta_TM_id
            FROM botApp_respusuariotamizaje ur 
            JOIN botApp_usuario us ON ur.id_manychat = us.id_manychat 
            WHERE respuesta_TM_id IN (1,2,3)
            GROUP BY edad, ur.respuesta_TM_id 
            ORDER BY edad ASC;"""  
        )
        resultados = cursor.fetchall()

    r_uno_edad_eti = "Menor de 50 años"
    r_dos_edad_eti = "Entre 50 y 69 años"
    r_tres_edad_eti = "Mayor a 69 años"
    opciones_anios = [r_uno_edad_eti, r_dos_edad_eti, r_tres_edad_eti]
    cantidades_si = [0, 0, 0]
    cantidades_no = [0, 0, 0]
    cantidades_no_recuerdo = [0,0,0]

    edad_min = 50
    edad_max = 69

    # Iteramos sobre los resultados
    for edad, cantidad, respuesta in resultados:
        index = 0 if edad < edad_min else (1 if edad <= edad_max else 2)
        
        if respuesta == 1:
            cantidades_si[index] += cantidad
        elif respuesta == 2:
            cantidades_no[index] += cantidad
        elif respuesta == 3:
            cantidades_no_recuerdo[index] += cantidad

    x = np.arange(len(opciones_anios))  
    width = 0.25  

    fig, ax = plt.subplots(figsize=[18, 8])
    rects1 = ax.bar(x - width, cantidades_si, width, label='Cantidad Sí', color = '#79addc')
    rects2 = ax.bar(x, cantidades_no, width, label='Cantidad No', color="#EFB0C9")
    rects3 = ax.bar(x + width, cantidades_no_recuerdo, width, label='Cantidad No Recuerdo', color="#A5F8CE")

    ax.set_xlabel("Rango de edad guía clínica")
    ax.set_ylabel("Número de Usuarias")
    ax.set_title("Mamografías por rango de Edad", pad=20)
    ax.set_xticks(x)
    ax.set_xticklabels(opciones_anios)
    plt.legend()

    # Etiquetas en las barras
    ax.bar_label(rects1, padding=3, color="black")
    ax.bar_label(rects2, padding=3, color="black")
    ax.bar_label(rects3, padding=3, color="black")

   # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def grafico_prev_salud_por_rango_edad():
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT us.edad, COUNT(*) as Cantidad, ur.respuesta_TM_id
            FROM botApp_usuario us JOIN  botApp_respusuariotamizaje ur ON us.id_manychat = ur.id_manychat
            WHERE ur.respuesta_TM_id IN(19,20,21)
            GROUP BY edad, respuesta_TM_id 
            ORDER BY edad ASC;"""  
        )
        resultados = cursor.fetchall()

    r_uno_edad_eti = "Menor de 50 años"
    r_dos_edad_eti = "Entre 50 y 69 años"
    r_tres_edad_eti = "Mayor a 69 años"
    opciones_anios = [r_uno_edad_eti, r_dos_edad_eti, r_tres_edad_eti]
    cantidades_fonasa = [0, 0, 0]
    cantidades_isapre = [0, 0, 0]
    cantidades_otro = [0, 0, 0]

    edad_min = 50
    edad_max = 69

    # Iteramos sobre los resultados
    for edad, cantidad, respuesta in resultados:
        index = 0 if edad < edad_min else (1 if edad <= edad_max else 2)
        
        if respuesta == 19:
            cantidades_fonasa[index] += cantidad
        elif respuesta == 20:
            cantidades_isapre[index] += cantidad
        elif respuesta == 21:
            cantidades_otro[index] += cantidad

    # Crear el gráfico
    plt.figure(figsize=[18, 8])
    plt.bar(opciones_anios, cantidades_fonasa, color="#a0b3a8", label="Cantidad Fonasa")
    plt.bar(opciones_anios, cantidades_isapre, color="#c6a78f", bottom=cantidades_fonasa, label="Cantidad Isapre")
    plt.bar(opciones_anios, cantidades_otro, color="#ecc8c9", bottom=np.array(cantidades_fonasa) + np.array(cantidades_isapre), label="Cantidad Otro")
    plt.xlabel("Rango de edad según guía clínica")
    plt.ylabel("Número de Usuarias")
    plt.title("Previsión de salud por rango de Edad", pad=20)
    plt.legend()

    # Agregar etiquetas para las barras de cantidades_fonasa
    for i, (edad, cantidad_fonasa) in enumerate(zip(opciones_anios, cantidades_fonasa)):
        if cantidad_fonasa > 0:
            plt.text(i, cantidad_fonasa / 2, str(cantidad_fonasa), ha='center', va='center', color='black', fontsize=10, fontweight='bold')

    # Agregar etiquetas para las barras de cantidades_isapre
    for i, (edad, cantidad_fonasa, cantidad_isapre) in enumerate(zip(opciones_anios, cantidades_fonasa, cantidades_isapre)):
        if cantidad_isapre > 0:
            plt.text(i, cantidad_fonasa + cantidad_isapre / 2, str(cantidad_isapre), ha='center', va='center', color='black', fontsize=10, fontweight='bold')

    # Agregar etiquetas para la barra de cantidades_otro
    for i, (edad, cantidad_fonasa, cantidad_isapre, cantidad_otro) in enumerate(zip(opciones_anios, cantidades_fonasa, cantidades_isapre, cantidades_otro)):
        if cantidad_otro > 0:
            plt.text(i, cantidad_fonasa + cantidad_isapre + cantidad_otro / 2, str(cantidad_otro), ha='center', va='center', color='black', fontsize=10, fontweight='bold')

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches="tight", dpi=300)
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_mamo_si_por_familiar_directo():
    
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT r.respuesta_TM_id, COUNT(DISTINCT r.id_manychat) AS cantidad_respuestas
            FROM botApp_respusuariotamizaje r
            WHERE r.respuesta_TM_id IN (25, 26, 27)  
            AND r.id_manychat IN (
            SELECT DISTINCT r2.id_manychat
            FROM botApp_respusuariotamizaje r2
            WHERE r2.respuesta_TM_id = 1  
            )   
            GROUP BY r.respuesta_TM_id;"""
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = OpcTamizaje.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.opc_respuesta_TM)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.opc_respuesta_TM} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['#79addc', '#EFB0C9', '#A5F8CE'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('Cantidad de usuarias que si se han realizado mamografías y tiene antecedentes familiares', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

def generar_grafico_mamo_no_por_familiar_directo():
    
    with connection.cursor() as cursor:
        cursor.execute(
            """SELECT r.respuesta_TM_id, COUNT(DISTINCT r.id_manychat) AS cantidad_respuestas
            FROM botApp_respusuariotamizaje r
            WHERE r.respuesta_TM_id IN (25, 26, 27)  
            AND r.id_manychat IN (
                SELECT DISTINCT r2.id_manychat
                FROM botApp_respusuariotamizaje r2
                WHERE r2.respuesta_TM_id = 2 
                )
            GROUP BY r.respuesta_TM_id;"""
        )
        resultados = cursor.fetchall()

    labels = []
    sizes = []
    counts = []

    for resultado in resultados:
        id_opc_respuesta, cantidad = resultado
        opcion_respuesta = OpcTamizaje.objects.get(id=id_opc_respuesta)
        labels.append(opcion_respuesta.opc_respuesta_TM)
        sizes.append(cantidad)
        counts.append(f"{opcion_respuesta.opc_respuesta_TM} - {cantidad}")

    # Configurar el gráfico circular
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(sizes, labels=None, autopct='%1.1f%%', startangle=90, colors=['#79addc', '#EFB0C9', '#A5F8CE'])
    
    # Configurar las etiquetas del gráfico
    ax.legend(wedges, counts, title="Respuestas", loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    
    # Mostrar el gráfico
    plt.title('Cantidad de usuarias que no se han realizado mamografías y tiene antecedentes familiares', pad=20)

    # Guardar la imagen en un buffer
    buffer = BytesIO()
    plt.savefig(buffer, format="png", bbox_inches='tight')
    buffer.seek(0)
    plt.close()

    # Convertir la imagen a base64
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return imagen_base64

@login_required
def reportes(request):
    data = {
        "imagen_base64_edad": generar_grafico_usuario_por_edad(),
        "imagen_base64_ingresos": generar_grafico_respuestas_por_dia(),
        "imagen_base64_ingresos_comuna": generar_grafico_ingresos_por_comuna(),
        "imagen_base64_pregunta1": generar_grafico_pregunta1(),
        "imagen_base64_pregunta2": generar_grafico_pregunta2(),
        "imagen_base64_pregunta3": generar_grafico_pregunta3(),
        "imagen_base64_pregunta4": generar_grafico_pregunta4(),
        "imagen_base64_pregunta5": generar_grafico_pregunta5(),
        "imagen_base64_pregunta6": generar_grafico_pregunta6(),
        "imagen_base64_referencias": generar_grafico_referencias(), 
        "imagen_base64_anios_nacimiento": generar_grafico_anio_nacimiento(),
        "imagen_base64_mamografia_si_por_edad":generar_grafico_mamografia_si_por_edad(),
        "imagen_base64_mamografia_no_por_edad":generar_grafico_mamografia_no_por_edad(),
        "imagen_base64_mamografia_por_edad_si_no": mamografia_por_edad_si_no(),
        "imagen_base64_tiempo_transc": generar_grafico_tiempo_trascurrido(),
        "imagen_base64_rango_edad": generar_grafico_por_rango_edad(),
        "imagen_base64_mamografia_si_no_rango_edad": mamografia_por_edad_si_no_rango_edad(),
        "imagen_base64_mamografia_si_no_rango_edad_agrupadas": mamografia_por_edad_si_no_rango_edad_agrupado(),
        "imagen_base64_grafico_prev_salud_por_rango_edad":grafico_prev_salud_por_rango_edad(),
        "imagen_base64_mamo_si_por_familiar_directo":generar_grafico_mamo_si_por_familiar_directo(),
        "imagen_base64_mamo_no_por_familiar_directo":generar_grafico_mamo_no_por_familiar_directo(),
        } 
    return render(request, "reportes.html", data)


# --------------------- Formulario WEB --------------------- #

"""
@login_required
def formulario(request):
    data = {
        "formUsuario": UsuarioForm(),
        "preguntas": PregTamizaje.objects.all(),
        "usuarios": User.objects.all(),
    }

    if request.method == "POST":
        form_usuario = UsuarioForm(request.POST)

        if form_usuario.is_valid():
            form_usuario.instance.id_usuario = request.POST.get("id_usuario")
            usuario = form_usuario.save()

            for pregunta in data["preguntas"]:
                respuesta = request.POST.get(f"pregunta_{pregunta.id}")
                opc_respuesta = opc_respuesta_TM(
                    id_pregunta=pregunta, OPC_Respuesta=respuesta
                )
                opc_respuesta.save()
                respuesta_usuario = RespuestaUsuario(
                    id_usuario=usuario,
                    id_pregunta=pregunta,
                    id_opc_respuesta=opc_respuesta,
                )
                respuesta_usuario.save()

            messages.success(request, "Datos guardados correctamente")
            form_usuario = UsuarioForm()
            return redirect(to="home")

        else:
            print(form_usuario.errors)
            messages.error(
                request,
                "La persona debe tener más de 18 años y haber nacido después de 1930.",
            )
            form_usuario = UsuarioForm()

    return render(request, "formulario.html", data)


"""
# --------------------- Preguntas --------------------- #

# Listar Preguntas
@login_required
def listarPreguntas(request):
    Preguntas = PregTamizaje.objects.all()
    data = {
        "preguntas": Preguntas,
    }
    return render(request, "preguntas/listarPreguntas.html", data)

# Modificar Pregunta
@login_required
def modificarPregunta(request, id):
    Preguntas = PregTamizaje.objects.get(id=id)
    data = {"form": PregTamizajeForm(instance=Preguntas)}

    if request.method == "POST":
        formulario = PregTamizajeForm(
            data=request.POST, instance=Preguntas, files=request.FILES
        )
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Modificado Correctamente")
            return redirect(to="listarPreguntas")
        data["form"] = formulario

    return render(request, "preguntas/modificarPreguntas.html", data)


# Eliminar Pregunta
@login_required
def eliminarPregunta(request, id):
    Preguntas = PregTamizaje.objects.get(id=id)
    Preguntas.delete()
    messages.success(request, "Eliminado Correctamente")
    return redirect(to="listarPreguntas")


# Crear Pregunta
@login_required
def crearPregunta(request):
    data = {"form": PregTamizajeForm()}

    if request.method == "POST":
        formulario = PregTamizajeForm(data=request.POST, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Pregunta Creada Correctamente")
        else:
            data["form"] = formulario

    return render(request, "preguntas/crearPreguntas.html", data)


# --------------------- Mensajes --------------------- #
@login_required
def homeMensajes(request):
    Mensajes = MensajeContenido.objects.all()
    data = {
        "mensajes": Mensajes,
    }
    return render(request, "mensajes/homeMensajes.html", data)

@login_required
def crearMensaje(request):
    data = {"form": MensajeContenidoForm()}

    if request.method == "POST":
        formulario = MensajeContenidoForm(data=request.POST, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Mensaje Creado Correctamente")
        else:
            data["form"] = formulario

    return render(request, "mensajes/crearMensajes.html", data)

@login_required
def modificarMensaje(request, id):
    Mensajes = MensajeContenido.objects.get(id=id)
    data = {"form": MensajeContenidoForm(instance=Mensajes)}

    if request.method == "POST":
        formulario = MensajeContenidoForm(
            data=request.POST, instance=Mensajes, files=request.FILES
        )
        if formulario.is_valid():
            formulario.save()
            messages.success(request, "Modificado Correctamente")
            return redirect(to="mensajesHome")
        data["form"] = formulario

    return render(request, "mensajes/modificarMensajes.html", data)

@login_required
def eliminarMensaje(request, id):
    Mensajes = MensajeContenido.objects.get(id=id)
    Mensajes.delete()
    messages.success(request, "Eliminado Correctamente")
    return redirect(to="mensajesHome")

# --------------------- Api --------------------- #

@login_required
def apiHome(request):
    return render(request, "api/apiHome.html")

#Usuario
class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer

#RespuestaUsuario
class UsuarioRespuestaViewSet(viewsets.ModelViewSet):
    queryset = RespUsuarioTamizaje.objects.all()
    serializer_class = UsuarioRespuestaSerializer

#MensajeContenido
class MensajeContenidoViewSet(viewsets.ModelViewSet):
    queryset = MensajeContenido.objects.all()
    serializer_class = MensajeContenidoSerializer

class UsuarioRespuestaAPIView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminUser]
    def get(self, request):
        respuestas = RespUsuarioTamizaje.objects.all()
        serializer = UsuarioRespuestaSerializer(respuestas, many=True)
        return Response(serializer.data)

class UsuarioAPIView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminUser]
    def get(self, request):
        usuarios = Usuario.objects.all()
        serializer = UsuarioSerializer(usuarios, many=True)
        return Response(serializer.data)
    
class MensajeContenidoAPIView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAdminUser]
    def get(self, request):
        mensajes = MensajeContenido.objects.all()
        serializer = MensajeContenidoSerializer(mensajes, many=True)
        return Response(serializer.data)
    
class ObtenerID(APIView):
    def get(self, request):
        # Obtener la fecha de hoy
        fecha_hoy = date.today()
        
        # Buscar un registro en la tabla que coincida con la fecha de hoy
        registro_hoy = MensajeContenido.objects.filter(fecha=fecha_hoy).first()
        
        if registro_hoy:
            # Si se encuentra un registro para la fecha de hoy, devolverlo
            return Response({'id': registro_hoy.id, 'texto': registro_hoy.texto, 'genero': registro_hoy.Genero_Usuario.OPC_Genero})
        else:
            # Si no se encuentra ningún registro para la fecha de hoy, devolver un código de error
            return Response({'error_code': '1'})
        
# --------------------- Api --------------------- #
def obtener_usuario(request, usuario_id):
    try:
        usuario = Usuario.objects.get(id=usuario_id)
        fecha_formateada = usuario.AnioNacimiento.strftime("%d/%m/%Y") if usuario.AnioNacimiento else None

        return JsonResponse({
            "nombre": usuario.nombre,
            "fecha_nacimiento": usuario.fecha_nacimiento,
            "AnioNacimiento": fecha_formateada  
        })
    except Usuario.DoesNotExist:
        return JsonResponse({"error": "Usuario no encontrado"}, status=404)


@csrf_exempt
def consultar_estado_pregunta(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido."}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Error al leer el JSON, asegúrate de que el formato es correcto."}, status=400)
    
    try:
        id_manychat = data["id_manychat"]
    except KeyError:
        return JsonResponse({"error": "El campo 'id_manychat' es requerido."}, status=400)

    # Verificar si el usuario existe en la BD
    try:
        usuario_model = Usuario.objects.get(id_manychat=id_manychat)
    except ObjectDoesNotExist:
        return JsonResponse({"respondido": "false"})

    # Inicializar respuesta como False
    respondido = False

    # Verificar el tipo de pregunta
    try:
        tipo_pregunta = data["tipo_pregunta"]
        nombre_pregunta = data["nombre_pregunta"]
    except KeyError:
        return JsonResponse({"error": "Los campos 'tipo_pregunta' y 'nombre_pregunta' son requeridos."}, status=400)

    if tipo_pregunta == "TM":
        pregunta_model = PregTamizaje.objects.filter(pregunta=nombre_pregunta).first()
        if pregunta_model:
            id_pregunta = pregunta_model.id
            opcion_respuesta_model = list(OpcTamizaje.objects.filter(id_pregunta=id_pregunta).values_list("id", flat=True))
            respuesta = list(RespUsuarioTamizaje.objects.filter(id_manychat=id_manychat, respuesta_TM_id__in=opcion_respuesta_model).    values_list("id", flat=True))
            respondido = len(respuesta) > 0

    """elif data["tipo_pregunta"] == "DS":
        pregunta_model = PregDeterSalud.objects.filter(pregunta_DS=data["nombre_pregunta"]).first()
        if pregunta_model:
            id_pregunta = pregunta_model.id
            opcion_respuesta_model = list(OpcDeterSalud.objects.filter(id_pregunta_DS=id_pregunta).values_list("id", flat=True))
            respuesta = list(RespDeterSalud.objects.filter(RutHash=rut_encriptado, respuesta_DS__in=opcion_respuesta_model).values_list("id", flat=True))
            respondido = len(respuesta) > 0

    elif data["tipo_pregunta"] == "FRM":
        pregunta_model = PregFactorRiesgoMod.objects.filter(pregunta_FRM=data["nombre_pregunta"]).first()
        if pregunta_model:
            id_pregunta = pregunta_model.id
            opcion_respuesta_model = list(OpcFactorRiesgoMod.objects.filter(id_pregunta_FRM=id_pregunta).values_list("id", flat=True))
            respuesta = list(RespUsuarioFactorRiesgoMod.objects.filter(RutHash=rut_encriptado, respuesta_FRM__in=opcion_respuesta_model).values_list("id", flat=True))
            respondido = len(respuesta) > 0

        # Verificación específica para peso y altura

            if "peso" in data["nombre_pregunta"].lower() or "altura" in data["nombre_pregunta"].lower():
            calculo_model = CalculoFRM.objects.filter(RutHash=rut_encriptado).first()
            if calculo_model:
                # Verificar si ambos valores (peso y altura) están presentes y son mayores que 0
                if calculo_model.peso_mod > 0 and calculo_model.altura_mod > 0:
                    respondido = True

    elif data["tipo_pregunta"] == "FRNM":
        pregunta_model = PregFactorRiesgoNoMod.objects.filter(pregunta_FRNM=data["nombre_pregunta"]).first()
        if pregunta_model:
            id_pregunta = pregunta_model.id
            opcion_respuesta_model = list(OpcFactorRiesgoNoMod.objects.filter(id_pregunta_FRNM=id_pregunta).values_list("id", flat=True))
            respuesta = list(RespUsuarioFactorRiesgoNoMod.objects.filter(RutHash=rut_encriptado, respuesta_FRNM__in=opcion_respuesta_model).values_list("id", flat=True))
            respondido = len(respuesta) > 0 """ 

    # Devolver la respuesta
    return JsonResponse({
        "respondido": "true" if respondido else "false"
    })

"""@csrf_exempt
def retorna_genero(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido."}, status=405)
    try:
        data = JSONParser().parse(request)
    except Exception:
        return JsonResponse({"error": "Error al leer el JSON, asegúrate de que el formato es correcto."}, status=400)

    # Validar que ManyChat envió el Rut
    if "Rut" not in data:
        return JsonResponse({"error": "El campo 'Rut' es obligatorio en la petición."}, status=400)

    # Encriptar el Rut para compararlo con RutHash
    rut_encriptado = generar_hash(data["Rut"])

    # Verificar si el usuario existe en la BD
    usuario_model = Usuario.objects.filter(RutHash=rut_encriptado).first()

    if usuario_model:
        pregunta_model = PregFactorRiesgoNoMod.objects.filter(pregunta_FRNM= "¿Cuál es tu género?").first()
        id_pregunta = pregunta_model.id
        opcion_respuesta_model = list(OpcFactorRiesgoNoMod.objects.filter(id_pregunta_FRNM=id_pregunta).values_list("id", flat=True))
        respuesta = RespUsuarioFactorRiesgoNoMod.objects.filter(RutHash=rut_encriptado, respuesta_FRNM__in=opcion_respuesta_model).first()
        #opcion = OpcFactorRiesgoNoMod.objects.filter(id=respuesta.respuesta_FRNM)
        return JsonResponse({"genero": respuesta.respuesta_FRNM.id})
    else:
        return JsonResponse({"error": "usuario no existe"})"""
    
@csrf_exempt
def verificar_usuario(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido."}, status=405)

    try:
        data = JSONParser().parse(request)
    except Exception as e:
        return JsonResponse({"error": f"Error al leer el JSON: {str(e)}"}, status=400)

    id_manychat = data["id_manychat"]
    
    try:
        # Verificar si el usuario existe en la BD
        usuario_model = Usuario.objects.filter(id_manychat=id_manychat).first()

        if usuario_model:
            return JsonResponse({"existe": "true"})
        else:
            return JsonResponse({"existe": "false"})

    except Exception as e:
        # Manejo de errores inesperados
        return JsonResponse({"error": f"Error interno del servidor: {str(e)}"}, status=500)

